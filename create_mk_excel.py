import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# ============================================================================
# KURIKULUM PROGRAM STUDI SARJANA TEKNIK KOMPUTER
# ============================================================================
# Target: 144-148 SKS Total
# Semester 8: HANYA 3 MK (Seminar Usulan, Seminar Hasil, Tugas Akhir)
# ============================================================================

def get_styles():
    styles = {
        'title_font': Font(bold=True, size=16, color="1F4E79", name="Calibri"),
        'subtitle_font': Font(bold=True, size=14, color="2E75B6", name="Calibri"),
        'header_font': Font(bold=True, color="FFFFFF", size=11, name="Calibri"),
        'content_font': Font(size=10, name="Calibri"),
        'bold_font': Font(bold=True, size=10, name="Calibri"),
        'header_fill_blue': PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid"),
        'univ_fill': PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid"),
        'fak_fill': PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
        'prodi_fill': PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid"),
        'total_fill': PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"),
        'thin_border': Border(
            left=Side(style='thin'), right=Side(style='thin'),
            top=Side(style='thin'), bottom=Side(style='thin')
        ),
        'thick_border': Border(
            left=Side(style='medium'), right=Side(style='medium'),
            top=Side(style='medium'), bottom=Side(style='medium')
        ),
    }
    return styles

def auto_adjust_column_width(ws, min_width=8, max_width=60):
    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        column = get_column_letter(col_idx)
        for row_idx in range(1, ws.max_row + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        adjusted_width = min(max(max_length * 1.2, min_width), max_width)
        ws.column_dimensions[column].width = adjusted_width

# ============================================================================
# DEFINISI MATA KULIAH - TARGET 146 SKS
# ============================================================================

# MK Universitas (dari mk_informatika.xlsx) - Total: 15 SKS
mk_universitas = [
    ("MKU01", "Agama I", 2, 1, "Universitas"),
    ("MKU02", "Pancasila", 2, 2, "Universitas"),
    ("MKU03", "Bahasa Indonesia", 2, 3, "Universitas"),
    ("MKU04", "Kewirausahaan", 3, 3, "Universitas"),  # 3 SKS, Semester 3 (sama dgn Bahasa Indonesia)
    ("MKU05", "Kewarganegaraan", 2, 4, "Universitas"),
    ("MKU06", "Bahasa Inggris", 2, 5, "Universitas"),
    ("MKU07", "Agama II", 2, 6, "Universitas"),
]

# MK Fakultas (dari mk_informatika.xlsx) - Total: 4 SKS
# Hanya 2 MK Fakultas yang ada di kolom "MK Fakultas" di sheet terakhir
mk_fakultas = [
    ("MKF01", "Dasar-Dasar Sains dan Technopreneurship", 2, 4, "Fakultas"),  # MK32, 2 SKS
    ("MKF02", "Bahasa Inggris Lanjut", 2, 6, "Fakultas"),  # MK47, 2 SKS
]

# ============================================================================
# MK Penciri Prodi Teknik Komputer (dari APTIKOM Tabel 9)
# Total Prodi: 122 SKS
# ============================================================================

mk_prodi = [
    # ========== SEMESTER 1 (Target: 21 SKS) ==========
    # Univ: 2, Fak: 1, Prodi: 18
    ("MK01", "Pengantar Teknik Komputer", 2, 1, "Prodi"),
    ("MK04", "Matematika Diskrit", 3, 1, "Prodi"),
    ("MK05", "Aljabar Linier", 3, 1, "Prodi"),
    ("MK08", "Fisika Dasar", 3, 1, "Prodi"),
    ("MK11", "Pemrograman Dasar", 4, 1, "Prodi"),
    ("MK22", "Elektronika", 3, 1, "Prodi"),
    # Prodi: 2+3+3+3+4+3 = 18 | Total: 21 SKS

    # ========== SEMESTER 2 (Target: 21 SKS) ==========
    # Univ: 2, Fak: 2, Prodi: 17
    ("MK02", "Kalkulus", 3, 2, "Prodi"),
    ("MK12", "Struktur Data", 3, 2, "Prodi"),
    ("MK17", "Arsitektur Komputer", 3, 2, "Prodi"),
    ("MK24", "Sistem Digital", 3, 2, "Prodi"),
    ("MK32", "Mikroprosesor", 3, 2, "Prodi"),
    ("MK37", "Komunikasi Data", 2, 2, "Prodi"),
    # Prodi: 3+3+3+3+3+2 = 17 | Total: 21 SKS

    # ========== SEMESTER 3 (Target: 21 SKS) ==========
    # Univ: 2, Fak: 0, Prodi: 19
    ("MK03", "Kalkulus Lanjut", 3, 3, "Prodi"),
    ("MK06", "Statistika dan Probabilitas", 3, 3, "Prodi"),
    ("MK09", "Basis Data", 3, 3, "Prodi"),
    ("MK16", "Organisasi Komputer", 3, 3, "Prodi"),
    ("MK23", "Rangkaian Listrik", 3, 3, "Prodi"),
    ("MK13", "Pemrograman Berorientasi Objek", 4, 3, "Prodi"),
    # Prodi: 3+3+3+3+3+4 = 19 | Total: 21 SKS

    # ========== SEMESTER 4 (Target: 21 SKS) ==========
    # Univ: 2, Fak: 2, Prodi: 17
    ("MK15", "Interaksi Manusia dan Komputer", 3, 4, "Prodi"),
    ("MK19", "Sistem Operasi", 3, 4, "Prodi"),
    ("MK25", "Konsep Embedded Systems", 3, 4, "Prodi"),
    ("MK34", "Jaringan Komputer", 3, 4, "Prodi"),
    ("MK14", "Pemrograman Web", 3, 4, "Prodi"),
    ("MK07", "Metode Numerik", 2, 4, "Prodi"),
    # Prodi: 3+3+3+3+3+2 = 17 | Total: 21 SKS

    # ========== SEMESTER 5 (Target: 22 SKS) ==========
    # Univ: 2, Fak: 3, Prodi: 17
    ("MK18", "Sistem Kendali", 3, 5, "Prodi"),
    ("MK26", "Perancangan Embedded Systems", 3, 5, "Prodi"),
    ("MK31", "Antarmuka Peripheral", 3, 5, "Prodi"),
    ("MK35", "Keamanan Jaringan Komputer", 3, 5, "Prodi"),
    ("MK39", "Sensor dan Teknologi", 3, 5, "Prodi"),
    ("MK21", "Analisa Kinerja Sistem", 2, 5, "Prodi"),
    # Prodi: 3+3+3+3+3+2 = 17 | Total: 22 SKS

    # ========== SEMESTER 6 (Target: 22 SKS) ==========
    # Univ: 2, Fak: 2, Prodi: 18
    ("MK10", "Rekayasa Perangkat Lunak", 3, 6, "Prodi"),
    ("MK27", "Sistem Akuisisi Data", 3, 6, "Prodi"),
    ("MK28", "Pengolahan Sinyal Digital", 3, 6, "Prodi"),
    ("MK30", "Sistem Cerdas", 3, 6, "Prodi"),
    ("MK33", "Mekatronika / PLC", 3, 6, "Prodi"),
    ("MK29", "Robotika", 3, 6, "Prodi"),
    # Prodi: 3+3+3+3+3+3 = 18 | Total: 22 SKS

    # ========== SEMESTER 7 (Target: 12 SKS) ==========
    # Univ: 2, Fak: 0, Prodi: 10
    ("MK20", "Sistem Waktu Nyata", 3, 7, "Prodi"),
    ("MK38", "Pemrosesan Paralel", 2, 7, "Prodi"),
    ("MK40", "Jaringan Sensor Nirkabel", 2, 7, "Prodi"),
    ("MK36", "Mobile Computing", 3, 7, "Prodi"),
    # Prodi: 3+2+2+3 = 10 | Total: 12 SKS

    # ========== SEMESTER 8 (Target: 10 SKS) ==========
    # HANYA 3 MK TUGAS AKHIR - TIDAK BOLEH ADA MK LAIN
    ("MKT01", "Seminar Usulan Tugas Akhir", 2, 8, "Prodi"),
    ("MKT02", "Seminar Hasil Tugas Akhir", 2, 8, "Prodi"),
    ("MKT03", "Tugas Akhir", 6, 8, "Prodi"),
    # Prodi: 2+2+6 = 10 | Total: 10 SKS
]

# Combine all courses
all_courses = mk_universitas + mk_fakultas + mk_prodi

# ============================================================================
# VALIDASI
# ============================================================================

print("=" * 70)
print("VALIDASI KURIKULUM TEKNIK KOMPUTER")
print("=" * 70)

total_sks_all = 0
for sem in range(1, 9):
    sem_courses = [c for c in all_courses if c[3] == sem]
    sem_sks = sum(c[2] for c in sem_courses)
    total_sks_all += sem_sks

    univ = [c for c in sem_courses if c[4] == "Universitas"]
    fak = [c for c in sem_courses if c[4] == "Fakultas"]
    prodi = [c for c in sem_courses if c[4] == "Prodi"]

    if sem <= 6:
        status = "OK" if 20 <= sem_sks <= 24 else "PERLU CEK"
    elif sem == 7:
        status = "OK" if 10 <= sem_sks <= 14 else "PERLU CEK"
    else:  # Semester 8
        status = "OK" if sem_sks == 10 and len(sem_courses) == 3 else "PERLU CEK"

    print(f"\nSemester {sem}: {sem_sks} SKS ({len(sem_courses)} MK) [{status}]")
    print(f"  - MK Universitas: {sum(c[2] for c in univ)} SKS ({len(univ)} MK)")
    print(f"  - MK Fakultas   : {sum(c[2] for c in fak)} SKS ({len(fak)} MK)")
    print(f"  - MK Prodi      : {sum(c[2] for c in prodi)} SKS ({len(prodi)} MK)")

    if sem == 8:
        print(f"  Mata Kuliah Semester 8:")
        for c in sem_courses:
            print(f"    - {c[1]} ({c[2]} SKS)")

print(f"\n{'='*70}")
print(f"TOTAL SKS: {total_sks_all}")
print(f"TOTAL MK : {len(all_courses)}")

# Validasi target
if 144 <= total_sks_all <= 148:
    print(f"STATUS: OK (dalam range 144-148 SKS)")
else:
    print(f"STATUS: PERLU CEK (target 144-148 SKS)")
print(f"{'='*70}")

# ============================================================================
# CREATE WORKBOOK
# ============================================================================

wb = Workbook()
styles = get_styles()

# ============================================================================
# SHEET 1: RINGKASAN KURIKULUM
# ============================================================================

ws1 = wb.active
ws1.title = "Ringkasan Kurikulum"

ws1.merge_cells('A1:G1')
ws1['A1'] = "SUSUNAN MATA KULIAH PROGRAM STUDI SARJANA TEKNIK KOMPUTER"
ws1['A1'].font = styles['title_font']
ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[1].height = 30

ws1.merge_cells('A2:G2')
ws1['A2'] = "Berdasarkan Panduan Kurikulum APTIKOM (Tabel 9) dan SN-DIKTI"
ws1['A2'].font = Font(italic=True, size=11, name="Calibri")
ws1['A2'].alignment = Alignment(horizontal='center', vertical='center')

ws1['A4'] = "RINGKASAN SKS PER SEMESTER"
ws1['A4'].font = styles['subtitle_font']

summary_headers = ["Semester", "MK Universitas", "MK Fakultas", "MK Prodi", "Total SKS", "Jumlah MK"]
ws1.append([])
ws1.append(summary_headers)
for col in range(1, 7):
    cell = ws1.cell(row=6, column=col)
    cell.font = styles['header_font']
    cell.fill = styles['header_fill_blue']
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = styles['thick_border']
ws1.row_dimensions[6].height = 25

semester_summary = []
for sem in range(1, 9):
    univ_sks = sum(c[2] for c in mk_universitas if c[3] == sem)
    fak_sks = sum(c[2] for c in mk_fakultas if c[3] == sem)
    prodi_sks = sum(c[2] for c in mk_prodi if c[3] == sem)
    total_sks = univ_sks + fak_sks + prodi_sks
    total_mk = sum(1 for c in all_courses if c[3] == sem)
    semester_summary.append((f"Semester {sem}", univ_sks, fak_sks, prodi_sks, total_sks, total_mk))

row_num = 7
for data in semester_summary:
    ws1.append(data)
    for col in range(1, 7):
        cell = ws1.cell(row=row_num, column=col)
        cell.border = styles['thin_border']
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.font = styles['content_font']
        if col == 2 and data[1] > 0:
            cell.fill = styles['univ_fill']
        elif col == 3 and data[2] > 0:
            cell.fill = styles['fak_fill']
        elif col == 4 and data[3] > 0:
            cell.fill = styles['prodi_fill']
    ws1.row_dimensions[row_num].height = 22
    row_num += 1

total_univ = sum(c[2] for c in mk_universitas)
total_fak = sum(c[2] for c in mk_fakultas)
total_prodi = sum(c[2] for c in mk_prodi)
total_all = total_univ + total_fak + total_prodi
total_mk = len(all_courses)

ws1.append(["TOTAL", total_univ, total_fak, total_prodi, total_all, total_mk])
for col in range(1, 7):
    cell = ws1.cell(row=row_num, column=col)
    cell.border = styles['thick_border']
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.font = styles['bold_font']
    cell.fill = styles['total_fill']
ws1.row_dimensions[row_num].height = 25

row_num += 2
ws1.cell(row=row_num, column=1, value="Keterangan:").font = styles['bold_font']
row_num += 1
ws1.cell(row=row_num, column=1, value="MK Universitas").fill = styles['univ_fill']
ws1.cell(row=row_num, column=1).border = styles['thin_border']
ws1.cell(row=row_num, column=2, value="Mata Kuliah Wajib Universitas")
row_num += 1
ws1.cell(row=row_num, column=1, value="MK Fakultas").fill = styles['fak_fill']
ws1.cell(row=row_num, column=1).border = styles['thin_border']
ws1.cell(row=row_num, column=2, value="Mata Kuliah Wajib Fakultas")
row_num += 1
ws1.cell(row=row_num, column=1, value="MK Prodi").fill = styles['prodi_fill']
ws1.cell(row=row_num, column=1).border = styles['thin_border']
ws1.cell(row=row_num, column=2, value="Mata Kuliah Penciri Program Studi (APTIKOM)")

ws1.column_dimensions['A'].width = 12
ws1.column_dimensions['B'].width = 15
ws1.column_dimensions['C'].width = 12
ws1.column_dimensions['D'].width = 12
ws1.column_dimensions['E'].width = 12
ws1.column_dimensions['F'].width = 12

# ============================================================================
# SHEET 2: DAFTAR LENGKAP MATA KULIAH
# ============================================================================

ws2 = wb.create_sheet("Daftar Mata Kuliah")

ws2.merge_cells('A1:F1')
ws2['A1'] = "DAFTAR LENGKAP MATA KULIAH"
ws2['A1'].font = styles['title_font']
ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 30

ws2.merge_cells('A2:F2')
ws2['A2'] = "PROGRAM STUDI SARJANA TEKNIK KOMPUTER"
ws2['A2'].font = styles['subtitle_font']
ws2['A2'].alignment = Alignment(horizontal='center', vertical='center')

headers = ["No", "Kode MK", "Nama Mata Kuliah", "SKS", "Semester", "Kategori"]
ws2.append([])
ws2.append(headers)
for col in range(1, 7):
    cell = ws2.cell(row=4, column=col)
    cell.font = styles['header_font']
    cell.fill = styles['header_fill_blue']
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = styles['thick_border']
ws2.row_dimensions[4].height = 25

sorted_courses = sorted(all_courses, key=lambda x: (x[3], x[0]))

row_num = 5
for i, course in enumerate(sorted_courses, 1):
    ws2.append([i, course[0], course[1], course[2], course[3], course[4]])
    for col in range(1, 7):
        cell = ws2.cell(row=row_num, column=col)
        cell.border = styles['thin_border']
        cell.font = styles['content_font']
        if col in [1, 2, 4, 5]:
            cell.alignment = Alignment(horizontal='center', vertical='center')
        else:
            cell.alignment = Alignment(horizontal='left', vertical='center')
        if course[4] == "Universitas":
            cell.fill = styles['univ_fill']
        elif course[4] == "Fakultas":
            cell.fill = styles['fak_fill']
        else:
            cell.fill = styles['prodi_fill']
    ws2.row_dimensions[row_num].height = 20
    row_num += 1

ws2.append(["", "", "TOTAL SKS", total_all, "", ""])
for col in range(1, 7):
    cell = ws2.cell(row=row_num, column=col)
    cell.border = styles['thick_border']
    cell.font = styles['bold_font']
    cell.fill = styles['total_fill']
    cell.alignment = Alignment(horizontal='center', vertical='center')

ws2.column_dimensions['A'].width = 6
ws2.column_dimensions['B'].width = 10
ws2.column_dimensions['C'].width = 45
ws2.column_dimensions['D'].width = 8
ws2.column_dimensions['E'].width = 10
ws2.column_dimensions['F'].width = 12
ws2.freeze_panes = 'A5'

# ============================================================================
# SHEET 3-10: DETAIL PER SEMESTER
# ============================================================================

semester_colors = [
    ("4472C4", "D6DCE5"),
    ("548235", "E2EFDA"),
    ("C65911", "FCE4D6"),
    ("7030A0", "E4DFEC"),
    ("2E75B6", "DDEBF7"),
    ("BF8F00", "FFF2CC"),
    ("375623", "C6EFCE"),
    ("833C0C", "F4B183"),
]

for sem in range(1, 9):
    ws = wb.create_sheet(f"Semester {sem}")
    header_color, row_color = semester_colors[sem-1]

    sem_courses = [c for c in all_courses if c[3] == sem]
    sem_sks = sum(c[2] for c in sem_courses)

    ws.merge_cells('A1:E1')
    ws['A1'] = f"SEMESTER {sem}"
    ws['A1'].font = Font(bold=True, size=18, color=header_color, name="Calibri")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 35

    ws.merge_cells('A2:E2')
    ws['A2'] = f"Total: {sem_sks} SKS | {len(sem_courses)} Mata Kuliah"
    ws['A2'].font = Font(bold=True, size=12, name="Calibri")
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 25

    headers = ["No", "Kode MK", "Nama Mata Kuliah", "SKS", "Kategori"]
    ws.append([])
    ws.append(headers)
    header_fill = PatternFill(start_color=header_color, end_color=header_color, fill_type="solid")
    for col in range(1, 6):
        cell = ws.cell(row=4, column=col)
        cell.font = styles['header_font']
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = styles['thick_border']
    ws.row_dimensions[4].height = 28

    row_num = 5
    category_order = {"Universitas": 1, "Fakultas": 2, "Prodi": 3}
    sem_courses_sorted = sorted(sem_courses, key=lambda x: (category_order[x[4]], x[0]))

    for i, course in enumerate(sem_courses_sorted, 1):
        ws.append([i, course[0], course[1], course[2], course[4]])
        for col in range(1, 6):
            cell = ws.cell(row=row_num, column=col)
            cell.border = styles['thin_border']
            cell.font = styles['content_font']
            if col in [1, 2, 4]:
                cell.alignment = Alignment(horizontal='center', vertical='center')
            else:
                cell.alignment = Alignment(horizontal='left', vertical='center')
            if course[4] == "Universitas":
                cell.fill = styles['univ_fill']
            elif course[4] == "Fakultas":
                cell.fill = styles['fak_fill']
            else:
                row_fill = PatternFill(start_color=row_color, end_color=row_color, fill_type="solid")
                cell.fill = row_fill
        ws.row_dimensions[row_num].height = 22
        row_num += 1

    ws.append(["", "", "TOTAL SKS", sem_sks, ""])
    for col in range(1, 6):
        cell = ws.cell(row=row_num, column=col)
        cell.border = styles['thick_border']
        cell.font = styles['bold_font']
        cell.fill = styles['total_fill']
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[row_num].height = 25

    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 10
    ws.column_dimensions['C'].width = 45
    ws.column_dimensions['D'].width = 8
    ws.column_dimensions['E'].width = 12
    ws.freeze_panes = 'A5'

# ============================================================================
# SHEET 11: STRUKTUR KURIKULUM (TABEL 9 STYLE)
# ============================================================================

ws_struktur = wb.create_sheet("Struktur Kurikulum")

ws_struktur.merge_cells('A1:K1')
ws_struktur['A1'] = "STRUKTUR KURIKULUM PROGRAM STUDI SARJANA TEKNIK KOMPUTER"
ws_struktur['A1'].font = styles['title_font']
ws_struktur['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws_struktur.row_dimensions[1].height = 30

ws_struktur.merge_cells('A2:K2')
ws_struktur['A2'] = "Distribusi Mata Kuliah per Semester (Format Tabel 9 APTIKOM)"
ws_struktur['A2'].font = styles['subtitle_font']
ws_struktur['A2'].alignment = Alignment(horizontal='center', vertical='center')

ws_struktur.append([])
headers = ["Kode MK", "Nama Mata Kuliah", "SKS", "1", "2", "3", "4", "5", "6", "7", "8"]
ws_struktur.append(headers)
for col in range(1, 12):
    cell = ws_struktur.cell(row=4, column=col)
    cell.font = styles['header_font']
    cell.fill = styles['header_fill_blue']
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = styles['thick_border']
ws_struktur.row_dimensions[4].height = 28

row_num = 5
for course in sorted(all_courses, key=lambda x: (x[3], x[0])):
    row_data = [course[0], course[1], course[2]]
    for sem in range(1, 9):
        if course[3] == sem:
            row_data.append("V")
        else:
            row_data.append("")

    ws_struktur.append(row_data)
    for col in range(1, 12):
        cell = ws_struktur.cell(row=row_num, column=col)
        cell.border = styles['thin_border']
        cell.font = styles['content_font']
        cell.alignment = Alignment(horizontal='center', vertical='center')
        if col == 2:
            cell.alignment = Alignment(horizontal='left', vertical='center')
        if course[4] == "Universitas":
            cell.fill = styles['univ_fill']
        elif course[4] == "Fakultas":
            cell.fill = styles['fak_fill']
        else:
            cell.fill = styles['prodi_fill']
    ws_struktur.row_dimensions[row_num].height = 20
    row_num += 1

total_row = ["", "TOTAL SKS", total_all]
for sem in range(1, 9):
    sem_total = sum(c[2] for c in all_courses if c[3] == sem)
    total_row.append(sem_total)

ws_struktur.append(total_row)
for col in range(1, 12):
    cell = ws_struktur.cell(row=row_num, column=col)
    cell.border = styles['thick_border']
    cell.font = styles['bold_font']
    cell.fill = styles['total_fill']
    cell.alignment = Alignment(horizontal='center', vertical='center')

ws_struktur.column_dimensions['A'].width = 10
ws_struktur.column_dimensions['B'].width = 45
ws_struktur.column_dimensions['C'].width = 6
for col in ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
    ws_struktur.column_dimensions[col].width = 5
ws_struktur.freeze_panes = 'A5'

# ============================================================================
# SAVE WORKBOOK
# ============================================================================

output_path = '/Users/akhiyarwaladi/Documents/prodi_sistem_komputer/Susunan_MK_Teknik_Komputer.xlsx'
wb.save(output_path)

print(f"\nFile Excel berhasil dibuat: {output_path}")
