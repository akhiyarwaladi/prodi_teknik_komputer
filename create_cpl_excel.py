import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.utils import get_column_letter

def auto_adjust_column_width(ws, min_width=10, max_width=100):
    """Auto-adjust column widths based on content"""
    for column_cells in ws.columns:
        max_length = 0
        column = column_cells[0].column_letter
        for cell in column_cells:
            try:
                if cell.value:
                    cell_length = len(str(cell.value))
                    if cell_length > max_length:
                        max_length = cell_length
            except:
                pass
        adjusted_width = min(max(max_length * 1.1, min_width), max_width)
        ws.column_dimensions[column].width = adjusted_width

# Create workbook
wb = Workbook()

# Define styles
header_font = Font(bold=True, color="FFFFFF", size=12, name="Calibri")
title_font = Font(bold=True, size=16, color="1F4E79", name="Calibri")
subtitle_font = Font(bold=True, size=14, color="2E75B6", name="Calibri")
content_font = Font(size=11, name="Calibri")

# Colors
header_fill_blue = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
header_fill_green = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
header_fill_orange = PatternFill(start_color="C65911", end_color="C65911", fill_type="solid")
header_fill_purple = PatternFill(start_color="7030A0", end_color="7030A0", fill_type="solid")
sikap_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
umum_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
khusus_fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
pengetahuan_fill = PatternFill(start_color="E4DFEC", end_color="E4DFEC", fill_type="solid")
alt_row_fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
# Warna untuk kolom Sumber
sndikti_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")  # Biru muda
aptikom_fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")  # Orange muda

# Border styles
thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)
thick_border = Border(
    left=Side(style='medium'),
    right=Side(style='medium'),
    top=Side(style='medium'),
    bottom=Side(style='medium')
)

# ==================== DATA CPL ====================

# CPL SIKAP (dari SN-DIKTI)
sikap_data = [
    ("S1", "Sikap", "Bertakwa kepada Tuhan Yang Maha Esa dan mampu menunjukkan sikap religius", "SN-DIKTI"),
    ("S2", "Sikap", "Menjunjung tinggi nilai kemanusiaan dalam menjalankan tugas berdasarkan agama, moral, dan etika", "SN-DIKTI"),
    ("S3", "Sikap", "Berkontribusi dalam peningkatan mutu kehidupan bermasyarakat, berbangsa, bernegara, dan kemajuan peradaban berdasarkan Pancasila", "SN-DIKTI"),
    ("S4", "Sikap", "Berperan sebagai warga negara yang bangga dan cinta tanah air, memiliki nasionalisme serta rasa tanggungjawab pada negara dan bangsa", "SN-DIKTI"),
    ("S5", "Sikap", "Menghargai keanekaragaman budaya, pandangan, agama, dan kepercayaan, serta pendapat atau temuan orisinal orang lain", "SN-DIKTI"),
    ("S6", "Sikap", "Bekerja sama dan memiliki kepekaan sosial serta kepedulian terhadap masyarakat dan lingkungan", "SN-DIKTI"),
    ("S7", "Sikap", "Taat hukum dan disiplin dalam kehidupan bermasyarakat dan bernegara", "SN-DIKTI"),
    ("S8", "Sikap", "Menginternalisasi nilai, norma, dan etika akademik", "SN-DIKTI"),
    ("S9", "Sikap", "Menunjukkan sikap bertanggungjawab atas pekerjaan di bidang keahliannya secara mandiri", "SN-DIKTI"),
    ("S10", "Sikap", "Menginternalisasi semangat kemandirian, kejuangan, dan kewirausahaan", "SN-DIKTI"),
]

# CPL KETERAMPILAN UMUM (dari SN-DIKTI untuk Sarjana)
umum_data = [
    ("KU1", "Keterampilan Umum", "Mampu menerapkan pemikiran logis, kritis, sistematis, dan inovatif dalam konteks pengembangan atau implementasi ilmu pengetahuan dan teknologi yang memperhatikan dan menerapkan nilai humaniora yang sesuai dengan bidang keahliannya", "SN-DIKTI"),
    ("KU2", "Keterampilan Umum", "Mampu menunjukkan kinerja mandiri, bermutu, dan terukur", "SN-DIKTI"),
    ("KU3", "Keterampilan Umum", "Mampu mengkaji implikasi pengembangan atau implementasi ilmu pengetahuan teknologi yang memperhatikan dan menerapkan nilai humaniora sesuai dengan keahliannya berdasarkan kaidah, tata cara dan etika ilmiah dalam rangka menghasilkan solusi, gagasan, desain atau kritik seni, menyusun deskripsi saintifik hasil kajiannya dalam bentuk skripsi atau laporan tugas akhir, dan mengunggahnya dalam laman perguruan tinggi", "SN-DIKTI"),
    ("KU4", "Keterampilan Umum", "Menyusun deskripsi saintifik hasil kajian tersebut di atas dalam bentuk skripsi atau laporan tugas akhir, dan mengunggahnya dalam laman perguruan tinggi", "SN-DIKTI"),
    ("KU5", "Keterampilan Umum", "Mampu mengambil keputusan secara tepat dalam konteks penyelesaian masalah di bidang keahliannya, berdasarkan hasil analisis informasi dan data", "SN-DIKTI"),
    ("KU6", "Keterampilan Umum", "Mampu memelihara dan mengembangkan jaringan kerja dengan pembimbing, kolega, sejawat baik di dalam maupun di luar lembaganya", "SN-DIKTI"),
    ("KU7", "Keterampilan Umum", "Mampu bertanggungjawab atas pencapaian hasil kerja kelompok dan melakukan supervisi dan evaluasi terhadap penyelesaian pekerjaan yang ditugaskan kepada pekerja yang berada di bawah tanggungjawabnya", "SN-DIKTI"),
    ("KU8", "Keterampilan Umum", "Mampu melakukan proses evaluasi diri terhadap kelompok kerja yang berada dibawah tanggung jawabnya, dan mampu mengelola pembelajaran secara mandiri", "SN-DIKTI"),
    ("KU9", "Keterampilan Umum", "Mampu mendokumentasikan, menyimpan, mengamankan, dan menemukan kembali data untuk menjamin kesahihan dan mencegah plagiasi", "SN-DIKTI"),
]

# CPL PENGETAHUAN (dari APTIKOM - Tabel 2)
pengetahuan_data = [
    ("P1", "Pengetahuan", "Mampu menjelaskan dan menerapkan konsep-konsep bidang teknik komputer, matematika dan statistika serta sains dasar untuk mengembangkan keterampilan berpikir analitis yang kuat melalui pembelajaran empiris dan eksperimen", "APTIKOM"),
    ("P2", "Pengetahuan", "Mampu menguasai dan menerapkan konsep-konsep bidang teknik komputer untuk menyelesaikan permasalahan pada dunia usaha dan dunia industri", "APTIKOM"),
]

# CPL KETERAMPILAN UMUM APTIKOM (tambahan dari APTIKOM - CPL03)
umum_aptikom_data = [
    ("KUA1", "Keterampilan Umum", "Mampu menelaah dan menyelesaikan permasalahan di bidang dunia usaha dan industri yang meliputi system sensor, jaringan sensor maupun Internet of Things (IoT), embedded systems dan akuisisi data dengan pemodelan, prototype maupun melalui simulasi komputer", "APTIKOM"),
]

# CPL KETERAMPILAN KHUSUS (dari APTIKOM - CPL04, CPL05)
khusus_data = [
    ("KK1", "Keterampilan Khusus", "Mampu menganalisis computing yang kompleks, merancang dan menerapkan inovasi perangkat sistem berbasis komputer yang meliputi system sensor, jaringan sensor maupun Internet of Things (IoT), embedded system dan akuisisi data untuk menghasilkan fungsi terbaru dengan kompleksitas yang lebih tinggi yang dibutuhkan oleh dunia usaha dan dunia industri", "APTIKOM"),
    ("KK2", "Keterampilan Khusus", "Mampu melakukan pemeliharaan dan pengujian sistem berbasis komputer yang memenuhi standar industri atau standar baku yang berlaku", "APTIKOM"),
]

# ==================== SHEET 1: RINGKASAN CPL ====================
ws1 = wb.active
ws1.title = "Ringkasan CPL"

# Title Section
ws1.merge_cells('A1:D1')
ws1['A1'] = "CAPAIAN PEMBELAJARAN LULUSAN (CPL)"
ws1['A1'].font = title_font
ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[1].height = 30

ws1.merge_cells('A2:D2')
ws1['A2'] = "PROGRAM STUDI SARJANA TEKNIK KOMPUTER"
ws1['A2'].font = subtitle_font
ws1['A2'].alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[2].height = 25

ws1.merge_cells('A3:D3')
ws1['A3'] = "Berdasarkan SN-DIKTI dan Panduan Kurikulum APTIKOM"
ws1['A3'].font = Font(italic=True, size=11, name="Calibri")
ws1['A3'].alignment = Alignment(horizontal='center', vertical='center')

# Empty row
ws1.append([])

# Headers
headers = ["Kode CPL", "Kategori", "Deskripsi CPL", "Sumber"]
ws1.append(headers)
header_row = 5
for col in range(1, 5):
    cell = ws1.cell(row=header_row, column=col)
    cell.font = header_font
    cell.fill = header_fill_blue
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thick_border
ws1.row_dimensions[header_row].height = 25

# Add data
row_num = 6

# Urutan: SIKAP → UMUM → PENGETAHUAN → KHUSUS

# 1. Add Sikap data (SN-DIKTI)
for i, data in enumerate(sikap_data):
    ws1.append(data)
    for col in range(1, 5):
        cell = ws1.cell(row=row_num, column=col)
        cell.border = thin_border
        cell.font = content_font
        cell.alignment = Alignment(vertical='center', wrap_text=True, horizontal='left' if col == 3 else 'center')
        if col == 2:
            cell.fill = sikap_fill
            cell.font = Font(bold=True, size=11, color="548235", name="Calibri")
        if col == 4:  # Kolom Sumber
            cell.fill = sndikti_fill
            cell.font = Font(bold=True, size=11, color="1F4E79", name="Calibri")
    ws1.row_dimensions[row_num].height = 35
    row_num += 1

# 2. Add Keterampilan Umum data (SN-DIKTI)
for i, data in enumerate(umum_data):
    ws1.append(data)
    for col in range(1, 5):
        cell = ws1.cell(row=row_num, column=col)
        cell.border = thin_border
        cell.font = content_font
        cell.alignment = Alignment(vertical='center', wrap_text=True, horizontal='left' if col == 3 else 'center')
        if col == 2:
            cell.fill = umum_fill
            cell.font = Font(bold=True, size=11, color="2E75B6", name="Calibri")
        if col == 4:  # Kolom Sumber
            cell.fill = sndikti_fill
            cell.font = Font(bold=True, size=11, color="1F4E79", name="Calibri")
    ws1.row_dimensions[row_num].height = 45
    row_num += 1

# 2b. Add Keterampilan Umum APTIKOM data
for i, data in enumerate(umum_aptikom_data):
    ws1.append(data)
    for col in range(1, 5):
        cell = ws1.cell(row=row_num, column=col)
        cell.border = thin_border
        cell.font = content_font
        cell.alignment = Alignment(vertical='center', wrap_text=True, horizontal='left' if col == 3 else 'center')
        if col == 2:
            cell.fill = umum_fill
            cell.font = Font(bold=True, size=11, color="2E75B6", name="Calibri")
        if col == 4:  # Kolom Sumber
            cell.fill = aptikom_fill
            cell.font = Font(bold=True, size=11, color="C65911", name="Calibri")
    ws1.row_dimensions[row_num].height = 50
    row_num += 1

# 3. Add Pengetahuan data (APTIKOM)
for i, data in enumerate(pengetahuan_data):
    ws1.append(data)
    for col in range(1, 5):
        cell = ws1.cell(row=row_num, column=col)
        cell.border = thin_border
        cell.font = content_font
        cell.alignment = Alignment(vertical='center', wrap_text=True, horizontal='left' if col == 3 else 'center')
        if col == 2:
            cell.fill = pengetahuan_fill
            cell.font = Font(bold=True, size=11, color="7030A0", name="Calibri")
        if col == 4:  # Kolom Sumber
            cell.fill = aptikom_fill
            cell.font = Font(bold=True, size=11, color="C65911", name="Calibri")
    ws1.row_dimensions[row_num].height = 50
    row_num += 1

# 4. Add Keterampilan Khusus data (APTIKOM)
for i, data in enumerate(khusus_data):
    ws1.append(data)
    for col in range(1, 5):
        cell = ws1.cell(row=row_num, column=col)
        cell.border = thin_border
        cell.font = content_font
        cell.alignment = Alignment(vertical='center', wrap_text=True, horizontal='left' if col == 3 else 'center')
        if col == 2:
            cell.fill = khusus_fill
            cell.font = Font(bold=True, size=11, color="C65911", name="Calibri")
        if col == 4:  # Kolom Sumber
            cell.fill = aptikom_fill
            cell.font = Font(bold=True, size=11, color="C65911", name="Calibri")
    ws1.row_dimensions[row_num].height = 50
    row_num += 1

# Set column widths for Sheet 1
ws1.column_dimensions['A'].width = 12
ws1.column_dimensions['B'].width = 22
ws1.column_dimensions['C'].width = 85
ws1.column_dimensions['D'].width = 12

# Freeze panes
ws1.freeze_panes = 'A6'

# ==================== SHEET 2: CPL SIKAP ====================
ws2 = wb.create_sheet("CPL Sikap")

ws2.merge_cells('A1:C1')
ws2['A1'] = "CPL SIKAP"
ws2['A1'].font = title_font
ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws2.row_dimensions[1].height = 30

ws2.merge_cells('A2:C2')
ws2['A2'] = "PROGRAM STUDI SARJANA TEKNIK KOMPUTER"
ws2['A2'].font = subtitle_font
ws2['A2'].alignment = Alignment(horizontal='center', vertical='center')

ws2.merge_cells('A3:C3')
ws2['A3'] = "Sumber: SN-DIKTI (Permendikbud No. 3 Tahun 2020)"
ws2['A3'].font = Font(italic=True, size=10, name="Calibri")
ws2['A3'].alignment = Alignment(horizontal='center')

ws2.append([])
headers2 = ["Kode", "Deskripsi Capaian Pembelajaran Sikap", "Indikator Pencapaian"]
ws2.append(headers2)
for col in range(1, 4):
    cell = ws2.cell(row=5, column=col)
    cell.font = header_font
    cell.fill = header_fill_green
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thick_border
ws2.row_dimensions[5].height = 30

row_num = 6
for data in sikap_data:
    ws2.append([data[0], data[2], ""])
    for col in range(1, 4):
        cell = ws2.cell(row=row_num, column=col)
        cell.border = thin_border
        cell.font = content_font
        cell.alignment = Alignment(vertical='center', wrap_text=True, horizontal='left' if col > 1 else 'center')
        if row_num % 2 == 0:
            cell.fill = sikap_fill
    ws2.row_dimensions[row_num].height = 40
    row_num += 1

ws2.column_dimensions['A'].width = 10
ws2.column_dimensions['B'].width = 75
ws2.column_dimensions['C'].width = 45
ws2.freeze_panes = 'A6'

# ==================== SHEET 3: CPL KETERAMPILAN UMUM ====================
ws3 = wb.create_sheet("CPL Keterampilan Umum")

ws3.merge_cells('A1:C1')
ws3['A1'] = "CPL KETERAMPILAN UMUM"
ws3['A1'].font = title_font
ws3['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws3.row_dimensions[1].height = 30

ws3.merge_cells('A2:C2')
ws3['A2'] = "PROGRAM STUDI SARJANA TEKNIK KOMPUTER"
ws3['A2'].font = subtitle_font
ws3['A2'].alignment = Alignment(horizontal='center', vertical='center')

ws3.merge_cells('A3:C3')
ws3['A3'] = "Sumber: SN-DIKTI (Permendikbud No. 3 Tahun 2020)"
ws3['A3'].font = Font(italic=True, size=10, name="Calibri")
ws3['A3'].alignment = Alignment(horizontal='center')

ws3.append([])
headers3 = ["Kode", "Deskripsi Capaian Pembelajaran Keterampilan Umum", "Indikator Pencapaian"]
ws3.append(headers3)
for col in range(1, 4):
    cell = ws3.cell(row=5, column=col)
    cell.font = header_font
    cell.fill = header_fill_blue
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thick_border
ws3.row_dimensions[5].height = 30

row_num = 6
for data in umum_data:
    ws3.append([data[0], data[2], ""])
    for col in range(1, 4):
        cell = ws3.cell(row=row_num, column=col)
        cell.border = thin_border
        cell.font = content_font
        cell.alignment = Alignment(vertical='center', wrap_text=True, horizontal='left' if col > 1 else 'center')
        if row_num % 2 == 0:
            cell.fill = umum_fill
    ws3.row_dimensions[row_num].height = 55
    row_num += 1

ws3.column_dimensions['A'].width = 10
ws3.column_dimensions['B'].width = 85
ws3.column_dimensions['C'].width = 45
ws3.freeze_panes = 'A6'

# ==================== SHEET 4: CPL APTIKOM (Pengetahuan + Keterampilan Khusus) ====================
ws4 = wb.create_sheet("CPL APTIKOM")

ws4.merge_cells('A1:D1')
ws4['A1'] = "CPL APTIKOM"
ws4['A1'].font = title_font
ws4['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws4.row_dimensions[1].height = 30

ws4.merge_cells('A2:D2')
ws4['A2'] = "PROGRAM STUDI SARJANA TEKNIK KOMPUTER"
ws4['A2'].font = subtitle_font
ws4['A2'].alignment = Alignment(horizontal='center', vertical='center')

ws4.merge_cells('A3:D3')
ws4['A3'] = "Sumber: Panduan Kurikulum APTIKOM (Tabel 2 - CPL Wajib Program Studi)"
ws4['A3'].font = Font(italic=True, size=10, name="Calibri")
ws4['A3'].alignment = Alignment(horizontal='center')

ws4.append([])
headers4 = ["Kode", "Kategori", "Deskripsi Capaian Pembelajaran", "Indikator Pencapaian"]
ws4.append(headers4)
for col in range(1, 5):
    cell = ws4.cell(row=5, column=col)
    cell.font = header_font
    cell.fill = header_fill_purple
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    cell.border = thick_border
ws4.row_dimensions[5].height = 30

row_num = 6

# Urutan APTIKOM: UMUM → PENGETAHUAN → KHUSUS

# 1. Add Keterampilan Umum APTIKOM data
for data in umum_aptikom_data:
    ws4.append([data[0], data[1], data[2], ""])
    for col in range(1, 5):
        cell = ws4.cell(row=row_num, column=col)
        cell.border = thin_border
        cell.font = content_font
        cell.alignment = Alignment(vertical='center', wrap_text=True, horizontal='left' if col == 3 else 'center')
        if col == 2:
            cell.fill = umum_fill
            cell.font = Font(bold=True, size=11, color="2E75B6", name="Calibri")
    ws4.row_dimensions[row_num].height = 55
    row_num += 1

# 2. Add Pengetahuan data from APTIKOM
for data in pengetahuan_data:
    ws4.append([data[0], data[1], data[2], ""])
    for col in range(1, 5):
        cell = ws4.cell(row=row_num, column=col)
        cell.border = thin_border
        cell.font = content_font
        cell.alignment = Alignment(vertical='center', wrap_text=True, horizontal='left' if col == 3 else 'center')
        if col == 2:
            cell.fill = pengetahuan_fill
            cell.font = Font(bold=True, size=11, color="7030A0", name="Calibri")
    ws4.row_dimensions[row_num].height = 55
    row_num += 1

# 3. Add Keterampilan Khusus data from APTIKOM
for data in khusus_data:
    ws4.append([data[0], data[1], data[2], ""])
    for col in range(1, 5):
        cell = ws4.cell(row=row_num, column=col)
        cell.border = thin_border
        cell.font = content_font
        cell.alignment = Alignment(vertical='center', wrap_text=True, horizontal='left' if col == 3 else 'center')
        if col == 2:
            cell.fill = khusus_fill
            cell.font = Font(bold=True, size=11, color="C65911", name="Calibri")
    ws4.row_dimensions[row_num].height = 65
    row_num += 1

ws4.column_dimensions['A'].width = 10
ws4.column_dimensions['B'].width = 22
ws4.column_dimensions['C'].width = 85
ws4.column_dimensions['D'].width = 45
ws4.freeze_panes = 'A6'

# Save workbook
output_path = '/Users/akhiyarwaladi/Documents/prodi_sistem_komputer/CPL_Teknik_Komputer.xlsx'
wb.save(output_path)
print(f"File Excel berhasil dibuat: {output_path}")
print("\nSheet yang tersedia:")
print("1. Ringkasan CPL - Semua CPL (urutan: Sikap → Umum → Pengetahuan → Khusus)")
print("2. CPL Sikap - 10 item dari SN-DIKTI")
print("3. CPL Keterampilan Umum - 9 item dari SN-DIKTI")
print("4. CPL APTIKOM - 5 item (1 Umum + 2 Pengetahuan + 2 Khusus)")
total_cpl = len(sikap_data) + len(pengetahuan_data) + len(umum_data) + len(umum_aptikom_data) + len(khusus_data)
print(f"\nTotal CPL: {total_cpl} item")
print("\nUrutan CPL: SIKAP → UMUM → PENGETAHUAN → KHUSUS")
